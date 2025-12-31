import argparse
import base64
import hashlib
import json
import os
import sys
from difflib import Differ
from pathlib import Path

import pandas as pd
from PIL import Image
from tqdm import tqdm

sys.path.append(str(Path(__file__).resolve().parent.parent))
from scoring.core.std_recog_score import compare


def diff_text(text1, text2):
    """计算两个文本的差异"""
    d = Differ()
    output = [(token[2:], token[0] if token[0] != ' ' else None) for token in d.compare(text2, text1)]
    return output


def diff_to_text(diff):
    """将diff结果转换为纯文本格式"""
    if not diff:
        return ""

    text_parts = []
    for text, status in diff:
        if status == '+':
            # 添加的文本用 [+text] 表示
            text_parts.append(f"[+{text}]")
        elif status == '-':
            # 删除的文本用 [-text] 表示
            text_parts.append(f"[-{text}]")
        else:
            # 未变化的文本正常显示
            text_parts.append(text)

    return ''.join(text_parts)


def diff_to_html(diff):
    """将diff结果转换为HTML格式的高亮文本"""
    if not diff:
        return ""

    html = []
    for text, status in diff:
        if status == '+':
            # 添加的文本用红色显示
            html.append(f'<span style="color: red; font-weight: bold;">{text}</span>')
        elif status == '-':
            # 删除的文本用绿色显示
            html.append(f'<span style="color: green; text-decoration: line-through;">{text}</span>')
        else:
            # 未变化的文本正常显示
            html.append(text)

    return ''.join(html)


def image_to_base64_html(image_path, max_width=200):
    """将图片转换为 Base64 以便在 HTML 中显示"""
    try:
        img = Image.open(image_path)
        w, h = img.size
        # 按比例缩放
        if w > max_width:
            ratio = max_width / w
            w = max_width
            h = int(h * ratio)

        with open(image_path, "rb") as img_file:
            base64_str = base64.b64encode(img_file.read()).decode("utf-8")

        # 获取图片格式
        img_format = img.format.lower() if img.format else 'png'
        if img_format == 'jpg':
            img_format = 'jpeg'

        return f'<img src="data:image/{img_format};base64,{base64_str}" width="{w}" height="{h}" style="max-width: {max_width}px;">'
    except Exception as e:
        return f"<span style='color: red;'>Error loading image: {str(e)}</span>"


def export_badcase(label_path, pred_path, image_dir, output_path, only_diff=False, output_format=None):
    """
    导出错例到文件（支持CSV、HTML、Excel格式）

    Args:
        label_path: 验证集标签文件路径（JSON格式）
        pred_path: 模型预估结果文件路径（JSON格式）
        image_dir: 验证集图片目录
        output_path: 输出文件路径
        only_diff: 是否只导出有差异的样本
        output_format: 输出格式 ('csv', 'html', 'xlsx')，如果为None则根据文件扩展名自动判断
    """
    # 根据文件扩展名自动判断格式
    if output_format is None:
        ext = os.path.splitext(output_path)[1].lower()
        if ext == '.html':
            output_format = 'html'
        elif ext == '.xlsx':
            output_format = 'xlsx'
        else:
            output_format = 'csv'

    # 读取标签和预测结果
    print(f"Loading labels from {label_path}...")
    with open(label_path, 'r', encoding='utf-8') as f:
        labels = json.loads(f.read())

    print(f"Loading predictions from {pred_path}...")
    with open(pred_path, 'r', encoding='utf-8') as f:
        preds = json.loads(f.read())

    print(f"Loaded {len(labels)} labels and {len(preds)} predictions.")

    # 处理每个图片
    results = []
    for img_name, info in tqdm(labels.items(), desc="Processing images"):
        gt = info['value'][0]
        from_image = info.get('from_image', '')  # 获取from_image字段

        try:
            pred = preds[img_name]['value'][0]
        except KeyError:
            print(f"Warning: No prediction found for {img_name}, skipping...")
            continue

        # 如果只导出有差异的样本，检查是否有差异
        if only_diff:
            score_depunc, _ = compare(pred=pred, label=gt, with_punctuation=False)
            if score_depunc == 1:
                continue

        # 构建图片路径
        img_path = os.path.join(image_dir, img_name)
        if not os.path.exists(img_path):
            print(f"Warning: File {img_path} does not exist, skipping...")
            continue

        # 计算diff
        diff = diff_text(pred, gt)
        diff_text_str = diff_to_text(diff)
        diff_html_str = diff_to_html(diff) if output_format == 'html' else None
        md5 = hashlib.md5(open(img_path, 'rb').read()).hexdigest()

        if output_format == 'html':
            # HTML格式：使用base64编码的图片
            image_html = image_to_base64_html(img_path)
            results.append(
                {
                    "image": image_html,
                    "diff": diff_html_str,
                    "image_path": img_path,
                    "ground_truth": gt,
                    "prediction": pred,
                    "md5": md5,
                    "from_image": from_image,
                }
            )
        elif output_format == 'xlsx':
            # Excel格式：保存图片路径和diff对象，后续会插入图片和富文本
            results.append(
                {
                    "image_path": img_path,
                    "diff": diff,  # 保存diff对象而不是文本，用于生成RichText
                    "diff_text": diff_text_str,  # 也保存文本作为备用
                    "ground_truth": gt,
                    "prediction": pred,
                    "md5": md5,
                    "from_image": from_image,
                }
            )
        else:
            # CSV格式：纯文本
            results.append({"image": img_path, "diff": diff_text_str, "md5": md5, "from_image": from_image})

    # 保存文件
    if output_format == 'html':
        export_to_html(results, output_path)
    elif output_format == 'xlsx':
        export_to_excel(results, output_path, image_dir)
    else:
        df = pd.DataFrame(results)
        df.to_csv(output_path, index=False, encoding='utf-8-sig')
        print(f"Exported {len(results)} cases to {output_path}")


def export_to_html(results, output_path):
    """导出为HTML格式"""
    html_content = """
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>Bad Cases Report</title>
        <style>
            body {{
                font-family: Arial, sans-serif;
                margin: 20px;
            }}
            table {{
                border-collapse: collapse;
                width: 100%;
                margin-top: 20px;
            }}
            th, td {{
                border: 1px solid #ddd;
                padding: 8px;
                text-align: left;
            }}
            th {{
                background-color: #4CAF50;
                color: white;
            }}
            tr:nth-child(even) {{
                background-color: #f2f2f2;
            }}
            img {{
                max-width: 200px;
                height: auto;
            }}
        </style>
    </head>
    <body>
        <h1>Bad Cases Report</h1>
        <p>Total cases: {total}</p>
        <table>
            <tr>
                <th>Image</th>
                <th>Diff</th>
                <th>Ground Truth</th>
                <th>Prediction</th>
                <th>MD5</th>
                <th>From Image</th>
            </tr>
    """.format(
        total=len(results)
    )

    for result in results:
        html_content += f"""
            <tr>
                <td>{result['image']}</td>
                <td>{result['diff']}</td>
                <td>{result.get('ground_truth', '')}</td>
                <td>{result.get('prediction', '')}</td>
                <td>{result.get('md5', '')}</td>
                <td>{result.get('from_image', '')}</td>
            </tr>
        """

    html_content += """
        </table>
    </body>
    </html>
    """

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
    print(f"Exported {len(results)} cases to {output_path}")


def diff_to_rich_text(diff):
    """将diff结果转换为openpyxl的RichText对象，支持富文本格式"""
    try:
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont
    except ImportError:
        return None

    if not diff:
        return CellRichText("")

    rich_text = CellRichText()
    for text, status in diff:
        if status == '+':
            # 添加的文本用红色加粗显示
            rich_text.append(TextBlock(InlineFont(color="FF0000", b=True), text))
        elif status == '-':
            # 删除的文本用绿色删除线显示
            rich_text.append(TextBlock(InlineFont(color="00AA00", strike=True), text))
        else:
            # 未变化的文本正常显示
            rich_text.append(text)

    return rich_text


def export_to_excel(results, output_path, image_dir):
    """导出为Excel格式（带图片和富文本diff）"""
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as OpenpyxlImage
        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Error: openpyxl is required for Excel export. Install it with: pip install openpyxl")
        print("Falling back to CSV format...")
        df = pd.DataFrame(
            [
                {
                    "image": r["image_path"],
                    "diff": r.get("diff_text", ""),
                    "md5": r.get("md5", ""),
                    "from_image": r.get("from_image", ""),
                }
                for r in results
            ]
        )
        csv_path = output_path.replace('.xlsx', '.csv')
        df.to_csv(csv_path, index=False, encoding='utf-8-sig')
        print(f"Exported {len(results)} cases to {csv_path}")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Bad Cases"

    # 设置表头
    headers = ["Image", "Diff", "Ground Truth", "Prediction", "MD5", "From Image"]
    ws.append(headers)

    # 设置列宽
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 30

    # 设置行高（用于显示图片）
    row_height = 100
    ws.row_dimensions[1].height = 20  # 表头行

    for idx, result in enumerate(tqdm(results, desc="Writing to Excel"), start=2):
        img_path = result['image_path']
        diff_obj = result.get('diff')  # diff对象
        diff_text = result.get('diff_text', '')  # 备用文本
        gt = result.get('ground_truth', '')
        pred = result.get('prediction', '')

        # 设置行高
        ws.row_dimensions[idx].height = row_height

        # 插入图片
        try:
            img = OpenpyxlImage(img_path)
            # 调整图片大小
            img.width = 200
            img.height = int(img.height * 200 / img.width)
            ws.add_image(img, f'A{idx}')
        except Exception as e:
            ws[f'A{idx}'].value = f"Error: {str(e)}"

        # 写入diff列，使用RichText格式
        diff_cell = ws[f'B{idx}']
        if diff_obj:
            rich_text = diff_to_rich_text(diff_obj)
            if rich_text:
                diff_cell.value = rich_text
            else:
                # 如果RichText不可用，使用普通文本
                diff_cell.value = diff_text
        else:
            diff_cell.value = diff_text

        # 写入其他列
        ws[f'C{idx}'] = gt
        ws[f'D{idx}'] = pred
        ws[f'E{idx}'] = result.get('md5', '')
        ws[f'F{idx}'] = result.get('from_image', '')

        # 设置文本换行
        alignment = Alignment(wrap_text=True, vertical='top')
        ws[f'B{idx}'].alignment = alignment
        ws[f'C{idx}'].alignment = alignment
        ws[f'D{idx}'].alignment = alignment
        ws[f'E{idx}'].alignment = alignment
        ws[f'F{idx}'].alignment = alignment

    wb.save(output_path)
    print(f"Exported {len(results)} cases to {output_path}")


def main():
    parser = argparse.ArgumentParser(description='Export bad cases to CSV/HTML/Excel')
    parser.add_argument('--label_path', type=str, required=True, help='验证集标签文件路径（JSON格式）')
    parser.add_argument('--pred_path', type=str, required=True, help='模型预估结果文件路径（JSON格式）')
    parser.add_argument('--image_dir', type=str, required=True, help='验证集图片目录')
    parser.add_argument('--output', type=str, required=True, help='输出文件路径（支持.csv, .html, .xlsx）')
    parser.add_argument(
        '--format',
        type=str,
        choices=['csv', 'html', 'xlsx'],
        default=None,
        help='输出格式（csv/html/xlsx），如果不指定则根据文件扩展名自动判断',
    )
    parser.add_argument('--only_diff', action='store_true', help='只导出有差异的样本')

    args = parser.parse_args()

    export_badcase(
        label_path=args.label_path,
        pred_path=args.pred_path,
        image_dir=args.image_dir,
        output_path=args.output,
        only_diff=args.only_diff,
        output_format=args.format,
    )


if __name__ == "__main__":
    main()
